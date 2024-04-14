from os import path, makedirs
import sys
from ttkbootstrap import Window, PhotoImage, Frame, Treeview, Scrollbar, Floodgauge, Style, Label, Button, SUCCESS, HEADINGS, LEFT, BOTH, CENTER, VERTICAL, RIGHT, Y, DETERMINATE, END, TOP, Toplevel, StringVar, Radiobutton, Canvas
from tkinter.messagebox import showinfo, askyesno, showwarning, showerror
from tkinter.filedialog import askopenfilenames
from typing import Tuple
from cytoflow import Tube, ImportOp, Experiment, ThresholdOp, DensityGateOp, FlowPeaksOp
from pandas import DataFrame, Categorical
from pandas.core.groupby import DataFrameGroupBy
from numpy import float64, ndarray, argmax, sort
from datetime import datetime
from csv import writer
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from pdfkit.configuration import Configuration
from pdfkit import configuration, from_string
from matplotlib.figure import Figure
from matplotlib.axes import Axes
from matplotlib.pyplot import subplots, show, close
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Constantes
_PROGRAM_DIRECTORY_NAME: str = "Tkinter experiment"
_USER_DIRECTORY_PATH: str = path.expanduser("~") # Esto obtiene el directorio de usuario
_PROGRAM_DIRECTORY_PATH: str = path.join(_USER_DIRECTORY_PATH, _PROGRAM_DIRECTORY_NAME) # Esto obtiene el directorio del programa
_THEME_NAME: str = "cyborg"
_APPLICATION_TITLE: str = "Tkinter experiment"
_BASE_ROUTE: str = getattr(sys, "_MEIPASS", path.abspath(path.dirname(__file__)))
_ICON_IMAGE_NAME: str = "tkinter_experiment_icon.png"
_ICON_IMAGE_PATH: str = path.join(_BASE_ROUTE, _ICON_IMAGE_NAME) # Esto obtiene la imagen de la pantalla de bienvenida
_MINIMUM_WINDOW_WIDTH: int = 800
_MINIMUM_WINDOW_HEIGHT: int = 600
# _BLACK = "black"#FIXME
_TREEVIEW_COLUMNS: tuple = ("file_name", "total_number_events", "number_cluster_events", "percentage_number_events_total", "mfi_cluster") # Definimos las columnas del treeview
_GREEN: str = "green"
_LIGHTGREEN: str = "lightgreen"
_WHITE: str = "white"
_MAXIMUM: str = "maximum"
_VALUE: str = "value"
_MASK: str = "mask"
_COLUMNS: str = "columns"
_VALUES: str = "values"
_TEXT: str = "text"
_XCHANNEL: str = "R1-A"
_YCHANNEL: str = "B8-A"
_SCALE: str = "log"
_CHANNEL: str = "B4-A"
_CLUSTER_NAME: str = "FlowPeaks"
_0DISABLED = "-disabled"
_MINIMUM_FLOODGAUGE_TOPLEVEL_WIDTH: int = 250
_MINIMUM_FLOODGAUGE_TOPLEVEL_HEIGHT: int = 75
_EXPORT_TOPLEVEL_TITLE: str = "Export to"
_MINIMUM_EXPORT_TOPLEVEL_WIDTH: int = 300
_MINIMUM_EXPORT_TOPLEVEL_HEIGHT: int = 200
_CSV_EXTENSION: str = ".csv"
_EXPORT_CSV_VALUE: str = f"{_EXPORT_TOPLEVEL_TITLE} {_CSV_EXTENSION}"
_REPORTS_DIRECTORY_NAME: str = "Reports"
_REPORTS_DIRECTORY_PATH: str = path.join(_PROGRAM_DIRECTORY_PATH, _REPORTS_DIRECTORY_NAME) # Esto obtiene el directorio de reportes
_W: str = "w"
_XLSX_EXTENSION: str = ".xlsx"
_EXPORT_XLSX_VALUE: str = f"{_EXPORT_TOPLEVEL_TITLE} {_XLSX_EXTENSION}"
_PDF_EXTENSION: str = ".pdf"
_EXPORT_PDF_VALUE: str = f"{_EXPORT_TOPLEVEL_TITLE} {_PDF_EXTENSION}"
_R: str = "r"
_TREEVIEW_FILE_NAME: str = "Treeview"
_TIMESTAMP: str = datetime.now().strftime("_%Y.%m.%d_%H.%M.%S")
_THIN: str = "thin"
_A: str = "A"
_B: str = "B"
_C: str = "C"
_D: str = "D"
_E: str = "E"
# Variables
_experiment_dictionary: dict = {}

# Función que genera el directorio del programa en el directorio de usuario
def generate_program_directory() -> None:
    """
    Función que genera el directorio del programa en el directorio de usuario
    """
    # Comprobamos si el directorio del programa existe y si no, lo crea
    if not path.exists(_PROGRAM_DIRECTORY_PATH):
        try:
            makedirs(_PROGRAM_DIRECTORY_PATH, exist_ok=True) # Esto crea el directorio del programa en el directorio de usuario
        except Exception as exception:
            showerror(title="Error", message=f"Error creating program directory:\n{exception}")

# Función que genera la ventana del programa
def generate_window() -> Window:
    """
    Función que genera la ventana del programa
    """
    # Generamos la ventana con un tema específico, un título, unas dimensiones mínimas y un ícono
    window: Window = Window(themename=_THEME_NAME)
    # window.withdraw()#FIXME
    window.title(_APPLICATION_TITLE)
    window.minsize(_MINIMUM_WINDOW_WIDTH, _MINIMUM_WINDOW_HEIGHT)
    icon_photo_image: PhotoImage = PhotoImage(file=_ICON_IMAGE_PATH)
    window.iconphoto(False, icon_photo_image)

    # Retornamos la ventana del programa
    return window

# Función que dimensiona y posiciona la ventana en la pantalla
def window_size_placement(window: Window) -> None:
    """
    Función que dimensiona y posiciona la ventana en la pantalla
    """
    # Obtiene las dimensiones de la pantalla
    screen_width: int = window.winfo_screenwidth()
    screen_height: int = window.winfo_screenheight()

    # Calcula la posición del centro
    position_top: int = int(screen_height / 2 - _MINIMUM_WINDOW_HEIGHT / 2)
    position_right: int = int(screen_width / 2 - _MINIMUM_WINDOW_WIDTH / 2)

    # Posiciona la ventana en el centro de la pantalla
    window.geometry(f"{_MINIMUM_WINDOW_WIDTH}x{_MINIMUM_WINDOW_HEIGHT}+{position_right}+{position_top}")

# FIXMEFunción que genera la pantalla de inicio
# def generate_splash_screen_toplevel() -> Toplevel:
#     """
#     Función que genera la pantalla de inicio
#     """
#     splash_screen_toplevel: Toplevel = Toplevel(background=_BLACK)
#     splash_screen_toplevel.overrideredirect(1) # Eliminamos la barra de título

#     return splash_screen_toplevel

# FIXMEFunción que dimensiona y posiciona el nivel superior de la pantalla de bienvenida
# def splash_screen_toplevel_size_placement(window: Window, splash_screen_toplevel: Toplevel) -> None:
#     """
#     Función que dimensiona y posiciona el nivel superior de la pantalla de bienvenida
#     """
#     SPLASH_SCREEN_TOPLEVEL_WIDTH: int = 632
#     SPLASH_SCREEN_TOPLEVEL_HEIGHT: int = 317

#     # Obtiene las dimensiones de la pantalla
#     screen_width: int = window.winfo_screenwidth()
#     screen_height: int = window.winfo_screenheight()

#     # Calcula la posición del centro
#     x: int = int(screen_width / 2 - SPLASH_SCREEN_TOPLEVEL_WIDTH / 2)
#     y: int = int(screen_height / 2 - SPLASH_SCREEN_TOPLEVEL_HEIGHT / 2)

#     # Posiciona el nivel superior de la pantalla de bienvenida en el centro de la pantalla
#     splash_screen_toplevel.geometry(f"{SPLASH_SCREEN_TOPLEVEL_WIDTH}x{SPLASH_SCREEN_TOPLEVEL_HEIGHT}+{x}+{y}")

# Función que genera el marco de los botones
def generate_buttons_frame(window: Window) -> Frame:
    """
    Función que genera el marco de los botones
    """
    # Creamos un frame en el que generaremos los botones
    buttons_frame: Frame = Frame(window)
    buttons_frame.place(relx=0, rely=0, relwidth=0.2, relheight=1)

    return buttons_frame

# Función que genera el marco de la tabla de datos
def generate_treeview_frame(window: Window) -> Frame:
    """
    Función que genera el marco de la tabla de datos
    """
    # Creamos un frame en el que generaremos el treeview
    treeview_frame: Frame = Frame(window)
    treeview_frame.place(relx=0.2, rely=0, relwidth=0.8, relheight=0.5)

    return treeview_frame

# Función que genera el marco de abajo
def generate_below_frame(window: Window) -> Frame:
    """
    Función que genera el marco de abajo
    """
    # Creamos un frame en el que generaremos el canvas y del botón de mostrar el canvas
    below_frame: Frame = Frame(window)
    below_frame.place(relx=0.2, rely=0.5, relwidth=0.8, relheight=0.5)

    return below_frame

# Función que genera el marco del canvas
def generate_canvas_frame(below_frame: Frame) -> Frame:
    """
    Función que genera el marco del canvas
    """
    # Creamos un frame en el que generaremos el canvas
    canvas_frame: Frame = Frame(below_frame)
    canvas_frame.place(relx=0, rely=0, relwidth=1, relheight=0.85)

    return canvas_frame

# Función que genera el marco del botón de mostrar el canvas
def generate_canvas_button_frame(below_frame: Frame) -> Frame:
    """
    Función que genera el marco del botón de mostrar el canvas
    """
    # Creamos un frame en el que generaremos el canvas y del botón de mostrar el canvas
    canvas_button_frame: Frame = Frame(below_frame)
    canvas_button_frame.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)

    return canvas_button_frame

# Función que genera la tabla con los datos
def generate_treeview(treeview_frame: Frame) -> Treeview:
    """
    Función que genera la tabla con los datos
    """
    # Creamos el treeview
    treeview: Treeview = Treeview(treeview_frame, bootstyle=SUCCESS, columns=_TREEVIEW_COLUMNS, show=HEADINGS) # NOTA: Para mostrar la columna #0, poner el atributo show=TREEHEADINGS
    treeview.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
    treeview.column(_TREEVIEW_COLUMNS[0], minwidth=152, width=152, anchor=CENTER)
    treeview.column(_TREEVIEW_COLUMNS[1], minwidth=110, width=110, anchor=CENTER)
    treeview.column(_TREEVIEW_COLUMNS[2], minwidth=125, width=125, anchor=CENTER)
    treeview.column(_TREEVIEW_COLUMNS[3], minwidth=150, width=150, anchor=CENTER)
    treeview.column(_TREEVIEW_COLUMNS[4], minwidth=75, width=75, anchor=CENTER)

    # Definimos las cabeceras
    treeview.heading(_TREEVIEW_COLUMNS[0], text="File name")
    treeview.heading(_TREEVIEW_COLUMNS[1], text="Total no. of events")
    treeview.heading(_TREEVIEW_COLUMNS[2], text="No. cluster events")
    treeview.heading(_TREEVIEW_COLUMNS[3], text="% no. of events over total")
    treeview.heading(_TREEVIEW_COLUMNS[4], text="MFI cluster")

    # Añadimos los scrollbars del treeview
    add_treeview_scrollbars(treeview_frame, treeview)

    return treeview

# Función que añade los scrollbars del treeview
def add_treeview_scrollbars(treeview_frame: Frame, treeview: Treeview) -> None:
    """
    Función que añade los scrollbars del treeview
    """
    custom_vertical_t_scrollbar_style: Style = Style()
    custom_vertical_t_scrollbar_style.configure("custom.Vertical.TScrollbar", background=_GREEN, troughcolor=_LIGHTGREEN, arrowcolor=_WHITE)

    vertical_scrollbar: Scrollbar = Scrollbar(treeview_frame, orient=VERTICAL, command=treeview.yview, style="custom.Vertical.TScrollbar")
    treeview.configure(yscrollcommand=vertical_scrollbar.set)
    vertical_scrollbar.pack(side=RIGHT, fill=Y)

# Función que genera los botones
def generate_buttons(window, buttons_frame: Frame, canvas_frame: Frame, canvas_button_frame: Frame, treeview: Treeview) -> None:
    """
    Función que genera los botones
    """
    # Ponemos un componente de relleno en la parte superior para poder centrar los botones verticalmente
    label1: Label = Label(buttons_frame)
    label1.pack(side=TOP, expand=True)
    
    # Creamos los botones
    # Creamos un estilo
    custom_t_button_style: Style = Style()
    custom_t_button_style.configure("custom.TButton", background=_GREEN, foreground=_LIGHTGREEN, font=("Helvetica", 12, "bold"))

    # Botón cargar archivos
    load_files_button: Button = Button(buttons_frame, text="Load files", command=lambda: select_fcs_files(treeview, window), style="custom.TButton")
    load_files_button.pack(side=TOP, pady=50)

    # Botón borrar
    delete_button: Button = Button(buttons_frame, text="Delete", command=lambda: delete_row(treeview, canvas_frame, canvas_button_frame), style="custom.TButton")
    delete_button.pack(side=TOP, pady=50)

    # Botón exportar
    export_button: Button = Button(buttons_frame, text="Export", command=lambda: export(window, treeview), style="custom.TButton")
    export_button.pack(side=TOP, pady=50)

    # Y también ponemos otro componente de relleno en la parte inferior para poder centrar los botones verticalmente
    label2: Label = Label(buttons_frame)
    label2.pack(side=TOP, expand=True)

# Función que selecciona archivos .fcs
def select_fcs_files(treeview: Treeview, window: Window) -> None:
    """
    Función que selecciona archivos .fcs
    """
    file_paths: tuple = askopenfilenames(initialdir=_BASE_ROUTE, filetypes=[("FCS files", "*.fcs")]) # Abre el cuadro de diálogo para seleccionar archivos
    if file_paths != "": # Si la lista de rutas de archivo no está vacía
        if treeview.get_children(): # Si el treeview contiene alguna línea
            treeview = delete_existing_elements_treeview(treeview) # Vaciamos el treeview
            _experiment_dictionary.clear() # Y también vaciamos el diccionario de experimentos
        
        floodgauge_toplevel: Toplevel = generate_floodgauge_toplevel(window)

        floodgauge_toplevel_size_placement(window, floodgauge_toplevel)

        custom_horizontal_t_floodgauge_style: Style = Style()
        custom_horizontal_t_floodgauge_style.configure("custom.Horizontal.TFloodgauge", background=_GREEN, troughcolor =_LIGHTGREEN, bordercolor=_GREEN)

        percentage: str = "0"
        floodgauge: Floodgauge = Floodgauge(floodgauge_toplevel, length=100, mode=DETERMINATE, style="custom.Horizontal.TFloodgauge", mask=update_floodgauge_mask(percentage), font=("Helvetica", 19, "bold"))
        floodgauge.pack(fill=BOTH, expand=True)

        # Comenzamos a procesar los ficheros
        process_files(floodgauge, file_paths, treeview, window, floodgauge_toplevel)

# Función que borra todos los elementos existentes en el treeview
def delete_existing_elements_treeview(treeview: Treeview) -> Treeview:
    """
    Función que borra todos los elementos existentes en el treeview
    """
    for i in treeview.get_children():
        treeview.delete(i)

    return treeview

# Función que genera el nivel superior
def generate_floodgauge_toplevel(window: Window) -> Toplevel:
    """
    Función que genera el nivel superior
    """
    # Generamos el nivel superior con un título y unas dimensiones mínimas
    floodgauge_toplevel: Toplevel = Toplevel(window)
    window.attributes(_0DISABLED, True) # Deshabilitamos la ventana principal
    floodgauge_toplevel.overrideredirect(1) # Eliminamos la barra de título
    floodgauge_toplevel.grab_set() # Hacemos que la ventana sea modal

    # Retornamos el nivel superior
    return floodgauge_toplevel

# Función que dimensiona y posiciona el nivel superior en la pantalla
def floodgauge_toplevel_size_placement(window: Window, floodgauge_toplevel: Toplevel) -> None:
    """
    Función que dimensiona y posiciona el nivel superior en la pantalla
    """
    # Obtenemos las dimensiones de la ventana principal
    window_width: int = window.winfo_width()
    window_height: int = window.winfo_height()

    # Obtenemos la posición de la ventana principal
    position_x: int = window.winfo_x()
    position_y: int = window.winfo_y()

    # Establecemos las dimensiones del toplevel
    floodgauge_toplevel.geometry(f"{_MINIMUM_FLOODGAUGE_TOPLEVEL_WIDTH}x{_MINIMUM_FLOODGAUGE_TOPLEVEL_HEIGHT}")

    # Calculamos la posición de la ventana secundaria
    floodgauge_toplevel.geometry("+%d+%d" % (position_x + (window_width - _MINIMUM_FLOODGAUGE_TOPLEVEL_WIDTH) / 2, position_y + (window_height - _MINIMUM_FLOODGAUGE_TOPLEVEL_HEIGHT) / 2))

# Función que actualiza la máscara de floodgauge
def update_floodgauge_mask(percentage: str) -> str:
    """
    Función que actualiza la máscara de floodgauge
    """
    return f"Processing\n{percentage}%"

# Función que procesa los archivos
def process_files(floodgauge: Floodgauge, file_paths: str, treeview: Treeview, window: Window, floodgauge_toplevel: Toplevel) -> None:
    """
    Función que procesa los archivos
    """
    number_file_paths: int = len(file_paths)
    floodgauge[_MAXIMUM] = number_file_paths

    for i, fcs_file in enumerate(file_paths):
        add_data_treeview(treeview, fcs_file) # Añadimos los datos al treeview

        floodgauge[_VALUE] = i + 1

        percentage: str = round((floodgauge[_VALUE] / number_file_paths) * 100)
        floodgauge[_MASK] = update_floodgauge_mask(percentage)
        
        floodgauge.update()
    
    window.attributes(_0DISABLED, False) # Habilitamos la ventana principal de nuevo
    
    floodgauge_toplevel.destroy()

    # Si el treeview contiene alguna línea, seleccionamos el primer elemento del treeview para mostrar el canvas
    if treeview.get_children():
        treeview.selection_set(treeview.get_children()[0])

# Función que añade los datos al treeview
def add_data_treeview(treeview: Treeview, fcs_file: str) -> Treeview:
    """
    Función que añade los datos al treeview
    """
    treeview.insert("", END, values=new_experiment(fcs_file))

    return treeview

# Función en la que aplicamos operaciones sobre el experimento y retornamos: el nombre del fichero, el nº de eventos total, el nº de eventos del cluster de interés, el % que representa el nº de eventos del cluster de interés sobre el total y la IMF del cluster de interés
def new_experiment(fcs_file: str) -> Tuple[str, int, int, str, str]:
    """
    Función en la que aplicamos operaciones sobre el experimento y retornamos: el nombre del fichero, el nº de eventos total, el nº de eventos del cluster de interés, el % que representa el nº de eventos del cluster de interés sobre el total y la IMF del cluster de interés
    """
    tube: Tube = Tube(file=fcs_file)

    import_op: ImportOp = ImportOp(tubes=[tube], channels={_XCHANNEL : _XCHANNEL, _YCHANNEL : _YCHANNEL, _CHANNEL : _CHANNEL})
    experiment_import: Experiment = import_op.apply()

    # Realizamos la operación Threshold sobre el experimento
    operation_name: str = "Threshold"
    threshold_op: ThresholdOp = ThresholdOp(name=operation_name, channel=_XCHANNEL, threshold=2000)
    experiment_threshold: Experiment = threshold_op.apply(experiment_import)
    experiment_threshold = experiment_threshold.query(operation_name)

    # Realizamos la operación DensityGate sobre el experimento
    operation_name: str = "DensityGate"
    density_gate_op: DensityGateOp = DensityGateOp(name=operation_name, xchannel=_XCHANNEL, xscale=_SCALE, ychannel=_YCHANNEL, yscale=_SCALE, keep=0.5)
    density_gate_op.estimate(experiment_threshold)
    experiment_density_gate: Experiment = density_gate_op.apply(experiment_threshold)
    experiment_density_gate = experiment_density_gate.query(operation_name)

    # Realizamos la operación de clustering FlowPeaks sobre el experimento
    flow_peaks_op: FlowPeaksOp = FlowPeaksOp(name=_CLUSTER_NAME, channels=[_XCHANNEL, _YCHANNEL], scale={_XCHANNEL : _SCALE, _YCHANNEL : _SCALE}, h0=3)
    flow_peaks_op.estimate(experiment_density_gate)
    experiment_flow_peaks: Experiment = flow_peaks_op.apply(experiment_density_gate)
    data_frame_experiment_flow_peaks_cluster_name: DataFrame = experiment_flow_peaks[[_CLUSTER_NAME]]
    data_frame_group_by_experiment_flow_peaks_cluster_name: DataFrameGroupBy = data_frame_experiment_flow_peaks_cluster_name.groupby(by=experiment_flow_peaks[_CLUSTER_NAME])
    argmax(data_frame_group_by_experiment_flow_peaks_cluster_name.count())

    # Asignamos a variables los datos que queremos retornar del experimento
    file_name: str = path.basename(fcs_file)

    data_frame_experiment_import: DataFrame = experiment_import.data
    total_number_events: int = data_frame_experiment_import.shape[0]

    data_frame_experiment_flow_peaks: DataFrame = experiment_flow_peaks.data
    number_events_cluster_interest: int = data_frame_experiment_flow_peaks.shape[0]

    percentage_represents_number_events_cluster_interest_total: str = "{:.0%}".format(round(number_events_cluster_interest / total_number_events, 2))

    mfi_cluster_interest: str = "{:.2f}".format(median_fluorescence_intensity_cluster_interest(experiment_flow_peaks))

    # Guardamos el experimento en un diccionario, tomando como clave el nombre del fichero .fcs
    _experiment_dictionary[file_name] = experiment_flow_peaks

    # Retornamos: el nombre del fichero, el nº de eventos total, el nº de eventos del cluster de interés, el % que representa el nº de eventos del cluster de interés sobre el total y la IMF del cluster de interés
    return (file_name, total_number_events, number_events_cluster_interest, percentage_represents_number_events_cluster_interest_total, mfi_cluster_interest)

# Función que calcula la Intensidad Mediana de Fluorescencia (IMF) sobre el cluster de interés
def median_fluorescence_intensity_cluster_interest(experiment_flow_peaks: Experiment) -> float64:
    """
    Función que calcula la Intensidad Mediana de Fluorescencia (IMF) sobre el cluster de interés
    """
    # Ordenamos los datos del experimento en el canal deseado
    sorted_data: ndarray = sort(experiment_flow_peaks[_CHANNEL])
    # Obtenemos el número total de datos
    total_number_data: int = len(sorted_data)
    
    # Si el número de datos es par
    if total_number_data % 2 == 0:
        return (sorted_data[total_number_data // 2 - 1] + sorted_data[total_number_data // 2]) / 2
    # Si el número de datos es impar
    else:
        return sorted_data[total_number_data // 2]

# Función que borra filas
def delete_row(treeview: Treeview, canvas_frame: Frame, canvas_button_frame: Frame) -> None:
    """
    Función que borra filas
    """
    selected_items: tuple = treeview.selection() # Esto devuelve una lista de los ID de las filas seleccionadas
    if selected_items: # Si hay alguna fila seleccionada
        confirm: bool = askyesno(title="Confirmation", message=f"Are you sure you want to delete {len(selected_items)} row{'' if len(selected_items) == 1 else 's'}?") # Mostramos un popup para confirmar o no que queremos borrar las filas
        if confirm: # Si hemos confirmado
            for selected_item in selected_items: # Borramos las filas seleccionadas
                valores = treeview.item(selected_item, _VALUES) # Obtenemos los valores de la fila
                experiment = valores[0] # Obtenemos el valor de la primera columna
                if experiment in _experiment_dictionary: # Comprobamos si la clave está en el diccionario
                    del _experiment_dictionary[experiment] # Borramos el elemento del diccionario
                treeview.delete(selected_item)
        if not treeview.get_children(): # Si el treeview se ha quedado vacío
            for widget in canvas_frame.winfo_children(): # Destruimos todos los componentes del marco del canvas
                widget.destroy()
            for widget in canvas_button_frame.winfo_children(): # Destruimos todos los componentes del marco del botón de mostrar el canvas
                widget.destroy()
    else: # En el caso de no haber ninguna fila seleccionada
        showwarning(title="Warning", message="No row selected") # Mostramos un cuadro de diálogo de advertencia

# Función que exporta el treeview a un fichero .csv
def export(window: Window, treeview: Treeview) -> None:
    """
    Función que exporta el treeview a un fichero .csv
    """
    if treeview.get_children(): # Si el treeview contiene alguna línea, se exportarán los datos del treeview, sino, aparecerá una advertencia diciendo que no hay datos que exportar del treeview
        export_toplevel: Toplevel = generate_export_toplevel(window)

        export_toplevel_size_placement(window, export_toplevel)

        options_frame: Frame = generate_export_options_frame(export_toplevel)
        accept_button_frame: Frame = generate_export_accept_button_frame(export_toplevel)

        string_var: StringVar = StringVar()

        custom_t_radiobutton_style: Style = Style()
        custom_t_radiobutton_style.configure("custom.TRadiobutton", font=("Helvetica", 12))

        csv_radiobutton = Radiobutton(options_frame, text=_EXPORT_CSV_VALUE, variable=string_var, value=_EXPORT_CSV_VALUE, style="custom.TRadiobutton")
        csv_radiobutton.pack(side=TOP, pady=(50, 0))

        xlsx_radiobutton = Radiobutton(options_frame, text=_EXPORT_XLSX_VALUE, variable=string_var, value=_EXPORT_XLSX_VALUE, style="custom.TRadiobutton")
        xlsx_radiobutton.pack(side=TOP)

        pdf_wkhtmltopdf_radiobutton = Radiobutton(options_frame, text=_EXPORT_PDF_VALUE, variable=string_var, value=_EXPORT_PDF_VALUE, style="custom.TRadiobutton")
        pdf_wkhtmltopdf_radiobutton.pack(side=TOP, pady=(0, 50))

        accept_button = Button(accept_button_frame, text="Accept", command=lambda: on_accept(string_var, treeview, export_toplevel, window))
        accept_button.pack()
    else:
        showwarning(title="Warning", message="There is no data in the treeview to export") # Mostramos un cuadro de diálogo de advertencia

# Función que genera el nivel superior
def generate_export_toplevel(window: Window) -> Toplevel:
    """
    Función que genera el nivel superior
    """
    # Generamos el nivel superior con un título y unas dimensiones mínimas
    export_toplevel: Toplevel = Toplevel(window)
    export_toplevel.title(_EXPORT_TOPLEVEL_TITLE)
    window.attributes(_0DISABLED, True) # Deshabilitamos la ventana principal
    export_toplevel.protocol("WM_DELETE_WINDOW", lambda: on_closing_export_toplevel(window, export_toplevel)) # Al cerrar el nivel superior, habilitamos la ventana principal de nuevo
    export_toplevel.resizable(False, False)
    export_toplevel.minsize(_MINIMUM_EXPORT_TOPLEVEL_WIDTH, _MINIMUM_EXPORT_TOPLEVEL_HEIGHT)
    export_toplevel.grab_set() # Hacemos que la ventana sea modal

    # Retornamos el nivel superior
    return export_toplevel

# Función que se va a ejecutar cuando cerremos el nivel superior de exportar
def on_closing_export_toplevel(window: Window, export_toplevel: Toplevel) -> None:
    """
    Función que se va a ejecutar cuando cerremos el nivel superior de exportar
    """
    window.attributes(_0DISABLED, False) # Habilitamos la ventana principal de nuevo
    
    export_toplevel.destroy()

# Función que dimensiona y posiciona el nivel superior en la pantalla
def export_toplevel_size_placement(window: Window, export_toplevel: Toplevel) -> None:
    """
    Función que dimensiona y posiciona el nivel superior en la pantalla
    """
    # Obtenemos las dimensiones de la ventana principal
    window_width: int = window.winfo_width()
    window_height: int = window.winfo_height()

    # Obtenemos la posición de la ventana principal
    position_x: int = window.winfo_x()
    position_y: int = window.winfo_y()

    # Establecemos las dimensiones del toplevel
    export_toplevel.geometry(f"{_MINIMUM_EXPORT_TOPLEVEL_WIDTH}x{_MINIMUM_EXPORT_TOPLEVEL_HEIGHT}")

    # Calculamos la posición de la ventana secundaria
    export_toplevel.geometry("+%d+%d" % (position_x + (window_width - _MINIMUM_EXPORT_TOPLEVEL_WIDTH) / 2, position_y + (window_height - _MINIMUM_EXPORT_TOPLEVEL_HEIGHT) / 2))

# Función que genera el marco de las opciones
def generate_export_options_frame(export_toplevel: Toplevel) -> Frame:
    """
    Función que genera el marco de las opciones
    """
    # Creamos un frame en el que generaremos las opciones
    options_frame: Frame = Frame(export_toplevel)
    options_frame.place(relx=0, rely=0, relwidth=1, relheight=0.75)

    return options_frame

# Función que genera el marco del botón aceptar
def generate_export_accept_button_frame(export_toplevel: Toplevel) -> Frame:
    """
    Función que genera el marco del botón aceptar
    """
    # Creamos un frame en el que generaremos el botón aceptar
    accept_button_frame: Frame = Frame(export_toplevel)
    accept_button_frame.place(relx=0, rely=0.75, relwidth=1, relheight=0.25)

    return accept_button_frame

# Función que ejecuta la correspondiente al aceptar según la opción seleccionada
def on_accept(string_var: StringVar, treeview: Treeview, export_toplevel: Toplevel, window: Window) -> None:
    """
    Función que ejecuta la correspondiente al aceptar según la opción seleccionada
    """
    option: str = string_var.get()

    if option == _EXPORT_CSV_VALUE:
        export_to_csv(treeview, window, export_toplevel)
    elif option == _EXPORT_XLSX_VALUE:
        export_to_xslx(treeview, window, export_toplevel)
    elif option == _EXPORT_PDF_VALUE:
        export_to_pdf(treeview, window, export_toplevel)
    else:
        export_toplevel.destroy()
        showwarning(title="Warning", message="No export method selected") # Mostramos un cuadro de diálogo de advertencia
        export(window, treeview)

# Función que exporta el treeview a un fichero .csv
def export_to_csv(treeview: Treeview, window: Window, export_toplevel: Toplevel) -> None:
    """
    Función que exporta el treeview a un fichero .csv
    """
    try:
        CSV_FILE_NAME: str = f"{_TREEVIEW_FILE_NAME}{_TIMESTAMP}{_CSV_EXTENSION}"
        generate_reports_directory()
        CSV_FILE_PATH: str = path.join(_REPORTS_DIRECTORY_PATH, CSV_FILE_NAME) # Esto une el directorio de reportes y "Treeview.csv" para formar una ruta completa donde se encuentra el archivo .csv
        with open(CSV_FILE_PATH, _W, newline="") as file:
            file_writer = writer(file)
            file_writer.writerow(treeview[_COLUMNS]) # Escribimos los nombres de las columnas
            for row_id in treeview.get_children():
                row: list = treeview.item(row_id)[_VALUES]
                file_writer.writerow(row)
        showinfo(title="Info", message=f"The treeview was exported to a {_CSV_EXTENSION} file in the reports directory")
    except Exception as exception:
        showerror(title="Error", message=f"Error when exporting the CSV:\n{str(exception)}")
    
    window.attributes(_0DISABLED, False) # Habilitamos la ventana principal de nuevo
    
    export_toplevel.destroy()

# Función que genera el directorio de reportes en el directorio del programa
def generate_reports_directory() -> None:
    """
    Función que genera el directorio de reportes en el directorio del programa
    """
    # Comprobamos si el directorio de reportes existe y si no, lo crea
    if not path.exists(_REPORTS_DIRECTORY_PATH):
        try:
            makedirs(_REPORTS_DIRECTORY_PATH, exist_ok=True) # Esto crea el directorio de reportes en el directorio del programa
        except Exception as exception:
            showerror(title="Error", message=f"Error creating reports directory:\n{exception}")

# Función que exporta el treeview a un fichero .xlsx
def export_to_xslx(treeview: Treeview, window: Window, export_toplevel: Toplevel) -> None:
    """
    Función que exporta el treeview a un fichero .xlsx
    """
    try:
        # Crea un nuevo libro de trabajo
        workbook: Workbook = Workbook()

        # Selecciona la hoja activa
        worksheet: Worksheet = workbook.active

        # Definir el estilo de relleno verde y letra negrita
        pattern_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        font = Font(bold=True)

        # Definir el estilo de borde
        border = Border(left=Side(style=_THIN), right=Side(style=_THIN), top=Side(style=_THIN), bottom=Side(style=_THIN))

        # Agregamos los datos de la tabla de datos
        treeview_header_text: list = [treeview.heading(column)[_TEXT] for column in treeview[_COLUMNS]] # Obtenemos en una lista los textos que se muestran en las cabeceras de la tabla de datos
        worksheet.append(treeview_header_text) # Añadimos los textos que se muestran en las cabeceras de la tabla de datos
        for cell in worksheet[1]: # Recorremos las celdas de la primera fila
            cell.fill = pattern_fill # Aplicamos un relleno 
            cell.font = font # Aplicamos una fuente
            cell.border = border # Aplicamos un borde
        
        for row_id in treeview.get_children():
            row: list = treeview.item(row_id)[_VALUES] # Obtenemos en una lista los datos de la tabla de datos
            worksheet.append(row) # Añadimos los datos de la tabla de datos

        # Modificamos el ancho de cada columna
        worksheet.column_dimensions[_A].width = 30.71 # 30.00
        worksheet.column_dimensions[_B].width = 17.57 # 16.86
        worksheet.column_dimensions[_C].width = 17.14 # 16.43
        worksheet.column_dimensions[_D].width = 23.86 # 23.14
        worksheet.column_dimensions[_E].width = 10.86 # 10.14

        # Centramos los datos de las columnas y aplicamos un borde a cada celda
        for column in [_A, _B, _C, _D, _E]:
            cell: Cell
            for cell in worksheet[column]:
                cell.alignment = Alignment(horizontal=CENTER)
                cell.border = border
        
        # Guardamos el libro de trabajo
        XLSX_FILE_NAME: str = f"{_TREEVIEW_FILE_NAME}{_TIMESTAMP}{_XLSX_EXTENSION}"
        generate_reports_directory()
        XLSX_FILE_PATH: str = path.join(_REPORTS_DIRECTORY_PATH, XLSX_FILE_NAME) # Esto une el directorio de reportes y "Treeview.xlsx" para formar una ruta completa donde se encuentra el archivo .xlsx
        workbook.save(XLSX_FILE_PATH)
        showinfo(title="Info", message=f"The treeview was exported to a {_XLSX_EXTENSION} file in the program directory")
    except Exception as exception:
        showerror(title="Error", message=f"Error when exporting the Excel:\n{str(exception)}")
    
    window.attributes(_0DISABLED, False) # Habilitamos la ventana principal de nuevo
    
    export_toplevel.destroy()

# Función que exporta el treeview a un fichero .pdf
def export_to_pdf(treeview: Treeview, window: Window, export_toplevel: Toplevel) -> None:
    """
    Función que exporta el treeview a un fichero .pdf
    """
    WKHTMLTOPDF_PROGRAM_PATH: str = path.join(_BASE_ROUTE, "wkhtmltopdf\\bin\\wkhtmltopdf.exe")
        
    if path.exists(WKHTMLTOPDF_PROGRAM_PATH): # Si existe la ruta del programa "wkhtmltopdf.exe" se podrá crear el PDF, sino saltará un aviso diciendo que no se ha encontrado la ruta
        try:
            # Generamos el PDF
            PDF_FILE_NAME: str = f"{_TREEVIEW_FILE_NAME}{_TIMESTAMP}{_PDF_EXTENSION}"
            generate_reports_directory()
            PDF_FILE_PATH: str = path.join(_REPORTS_DIRECTORY_PATH, PDF_FILE_NAME) # Esto une el directorio de reportes y "Treeview.pdf" para formar una ruta completa donde se encuentra el archivo .pdf
            PDFKIT_CONFIGURATION: Configuration = configuration(wkhtmltopdf=WKHTMLTOPDF_PROGRAM_PATH) # Configuración de pdfkit
            from_string(html_content(treeview), PDF_FILE_PATH, configuration=PDFKIT_CONFIGURATION)
            showinfo(title="Info", message=f"The treeview was exported to a {_PDF_EXTENSION} file in the program directory")
        except Exception as exception:
            showerror(title="Error", message=f"Error when exporting the PDF:\n{str(exception)}")
    else:
        showerror(title="Error", message="The program path \"wkhtmltopdf.exe\" could not be found. Cannot create PDF") # Mostramos un cuadro de diálogo de error

    window.attributes(_0DISABLED, False) # Habilitamos la ventana principal de nuevo
        
    export_toplevel.destroy()

# Función que obtiene el contenido HTML, que será utilizado para generar el PDF
def html_content(treeview: Treeview) -> str:
    """
    Función que obtiene el contenido HTML, que será utilizado para generar el PDF
    """
    # Leemos la plantilla HTML y la almacenamos en una variable
    HTML_TEMPLATE_PATH: str = path.join(_BASE_ROUTE, "template.html")
    html_content: str = ""
    if path.exists(HTML_TEMPLATE_PATH):
        with open(HTML_TEMPLATE_PATH, _R) as file:
            html_content += file.read()
    
    # Obtenemos en una lista los textos que se muestran en las cabeceras de la tabla de datos
    treeview_header_text: list = [treeview.heading(column)[_TEXT] for column in treeview[_COLUMNS]]
    
    th: str = ""
    for header_text in treeview_header_text:
        th += f"""
        {"".join([f"<th>{header_text}</th>"])}
        """

    # Obtenemos los datos del treeview
    treeview_data: list = []
    for row in treeview.get_children():
        treeview_data.append(treeview.item(row)[_VALUES])
    
    tr: str = ""
    for row in treeview_data:
        tr += f"""
        <tr>
            {"".join([f"<td>{row[i]}</td>" for i in range(len(_TREEVIEW_COLUMNS))])}
        </tr>
        """

    # Sustituimos los elementos que hay entre llaves en la plantilla HTML por los datos que necesitemos
    html_content = html_content.replace("{title}", _TREEVIEW_FILE_NAME)
    html_content = html_content.replace("{h1}", _TREEVIEW_FILE_NAME)
    html_content = html_content.replace("{th}", th)
    html_content = html_content.replace("{tr}", tr)

    return html_content

# Función que muestra el canvas
def show_canvas_selected_row_treeview(canvas_frame: Frame, canvas_button_frame: Frame, treeview: Treeview) -> None:
    """
    Función que muestra el canvas
    """
    selected_items: tuple = treeview.selection() # Esto devuelve una lista de los ID de las filas seleccionadas
    if len(selected_items) == 1:
        # Obtenemos el valor de la 1ª columna
        column_value: str = treeview.item(selected_items[0])[_VALUES][0]

        # Una vez realizadas las operaciones, pintamos la gráfica de puntos en la ventana del programa
        generate_canvas(canvas_frame, column_value)

        # Y generamos el botón por si queremos visualizar la gráfica en una ventana a parte
        generate_canvas_button(canvas_button_frame, canvas_frame, column_value)
    else:
        for widget in canvas_frame.winfo_children(): # Destruimos todos los componentes del marco del canvas
            widget.destroy()
        for widget in canvas_button_frame.winfo_children(): # Destruimos todos los componentes del marco del botón de mostrar el canvas
            widget.destroy()

# Función que pinta los eventos del cluster en una gráfica en la ventana del programa
def generate_canvas(canvas_frame: Frame, column_value: str, clic_view_graph_window_button: bool=False) -> None:
    """
    Función que pinta los eventos del cluster en una gráfica en la ventana del programa
    """
    for widget in canvas_frame.winfo_children(): # Destruimos todos los componentes del marco del canvas
        widget.destroy()
    
    experiment: Experiment = _experiment_dictionary.get(column_value)
    data_frame_experiment: DataFrame = experiment.data

    # Dibujamos la gráfica de puntos
    figure: Figure
    axes: Axes
    figure, axes = subplots()

    # Dibujamos los puntos, diferenciando los clusters por color
    clusters: Categorical = data_frame_experiment[_CLUSTER_NAME].unique()
    for cluster in clusters:
        data_frame_cluster: DataFrame = data_frame_experiment[data_frame_experiment[_CLUSTER_NAME] == cluster]
        axes.scatter(data_frame_cluster[_XCHANNEL], data_frame_cluster[_YCHANNEL], label=f"{_CLUSTER_NAME} {cluster}", s=5, color=_GREEN)

    # Mostramos la leyenda
    axes.legend()
    
    # Crear el canvas de tkinter y añadir la figura de matplotlib
    figure_canvas_tk_agg: FigureCanvasTkAgg = FigureCanvasTkAgg(figure, master=canvas_frame)
    figure_canvas_tk_agg.draw()
    canvas: Canvas = figure_canvas_tk_agg.get_tk_widget()
    canvas.pack(fill=BOTH, expand=True, padx=100, pady=(20, 5))

    # Si hemos pulsado el botón de ver gráfica en ventana
    if clic_view_graph_window_button:
        show()
    
    # Cerramos la figura para que el programa no se quede pillado al cerrarlo
    close(fig=figure)

# Función que genera el botón de ver la gráfica en ventana
def generate_canvas_button(canvas_button_frame: Frame, canvas_frame: Frame, column_value: str) -> None:
    """
    Función que genera el botón de ver la gráfica en ventana
    """
    for widget in canvas_button_frame.winfo_children(): # Destruimos todos los componentes del marco del botón de mostrar el canvas
        widget.destroy()
    
    # Creamos el botón ver gráfica en ventana
    view_graph_window_button: Button = Button(canvas_button_frame, text="View graph in window", command=lambda: show_graph_window(canvas_frame, column_value, True))
    view_graph_window_button.pack()

# Función que muestra la gráfica en ventana
def show_graph_window(canvas_frame: Frame, column_value: str, clic_view_graph_window_button: bool) -> None:
    """
    Función que muestra la gráfica en ventana
    """
    # Generamos el canvas y lo mostramos en una ventana
    generate_canvas(canvas_frame, column_value, clic_view_graph_window_button)

# Si el nombre del módulo es igual a __main__ se ejecutará el código (esto se hace por si queremos utilizar este código como un módulo, y no queremos que ejecute el código del main)
if __name__ == "__main__":
    # Generamos el directorio del programa en el directorio de usuario
    generate_program_directory()

    # Generamos la ventana del programa
    window: Window = generate_window()
    
    # Dimensionamos y posicionamos la ventana en la pantalla
    window_size_placement(window)

    #FIXME
    # # Generamos la pantalla de bienvenida
    # splash_screen_toplevel: Toplevel = generate_splash_screen_toplevel()
    
    # # Dimensionamos y posicionamos el nivel superior de la pantalla de bienvenida
    # splash_screen_toplevel_size_placement(window, splash_screen_toplevel)

    # # Cargar una imagen para la pantalla de bienvenida
    # SPLASH_SCREEN_IMAGE_NAME: str = "tkinter_ttkbootstrap_splash_screen.png"
    # SPLASH_SCREEN_IMAGE_PATH: str = path.join(_BASE_ROUTE, SPLASH_SCREEN_IMAGE_NAME) # Esto obtiene la imagen de la pantalla de bienvenida
    # splash_screen_photo_image: PhotoImage = PhotoImage(file=SPLASH_SCREEN_IMAGE_PATH)
    # label = Label(splash_screen_toplevel, image=splash_screen_photo_image)
    # label.pack()
    # splash_screen_toplevel.update()

    # MILLISECONDS: int = 5000
    # window.after(MILLISECONDS, window.deiconify)  # Muestra la ventana principal después de 100ms
    # window.after(MILLISECONDS, splash_screen_toplevel.destroy)  # Destruye la pantalla de inicio después de 100ms
    #FIXME

    # El estado de la ventana será maximizado
    # window.state("zoomed")

    # Generamos los marcos de la ventana del programa
    buttons_frame: Frame = generate_buttons_frame(window) # Generamos el marco de los botones
    treeview_frame: Frame = generate_treeview_frame(window) # Generamos el marco de la tabla de datos
    below_frame: Frame = generate_below_frame(window) # Generamos el marco del botón del canvas
    canvas_frame: Frame = generate_canvas_frame(below_frame) # Generamos el marco del canvas
    canvas_button_frame: Frame = generate_canvas_button_frame(below_frame) # Generamos el marco del botón del canvas

    # Generamos la tabla de datos
    treeview: Treeview = generate_treeview(treeview_frame)

    # Generamos los botones
    generate_buttons(window, buttons_frame, canvas_frame, canvas_button_frame, treeview)

    # Mostramos el canvas cada vez que seleccionamos una fila del treeview una vez que se ha cargado todo
    treeview.bind("<<TreeviewSelect>>", lambda event: show_canvas_selected_row_treeview(canvas_frame, canvas_button_frame, treeview))

    # Generamos el hilo que genera la ventana del programa
    window.mainloop()
